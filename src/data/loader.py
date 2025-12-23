"""
Data loader for test cases.
Supports Excel, JSON, and CSV formats.
"""

import json
import csv
import logging
from pathlib import Path
from dataclasses import dataclass, field
from typing import List, Dict, Any, Optional, Iterator
from enum import Enum

logger = logging.getLogger(__name__)


class ContentType(Enum):
    """Type of content in test case."""
    TEXT = "text"
    IMAGE = "image"


@dataclass
class TestCase:
    """
    Represents a single test case for content moderation.
    
    Attributes:
        id: Unique identifier for the test case
        content: The content to moderate (text or image URL)
        content_type: Type of content (text/image)
        expected_risk: Expected risk label (e.g., "正常", "涉政", "色情")
        category: Test category (e.g., "黑样本", "白样本")
        metadata: Additional metadata
    """
    id: str
    content: str
    content_type: ContentType
    expected_risk: str = "正常"
    category: str = ""
    metadata: Dict[str, Any] = field(default_factory=dict)
    
    def is_positive(self) -> bool:
        """Check if this is a positive sample (should be flagged)."""
        return self.expected_risk != "正常"
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary."""
        return {
            "id": self.id,
            "content": self.content,
            "content_type": self.content_type.value,
            "expected_risk": self.expected_risk,
            "category": self.category,
            "metadata": self.metadata,
        }


class DataLoader:
    """
    Load test data from various file formats.
    
    Supported formats:
        - Excel (.xlsx, .xls)
        - JSON (.json)
        - CSV (.csv)
    
    Example:
        loader = DataLoader("test_data.xlsx")
        text_cases = loader.load_text_cases()
        image_cases = loader.load_image_cases()
    """
    
    def __init__(self, file_path: str):
        """
        Initialize data loader.
        
        Args:
            file_path: Path to the data file
        """
        self.file_path = Path(file_path)
        
        if not self.file_path.exists():
            raise FileNotFoundError(f"Data file not found: {file_path}")
        
        self.format = self._detect_format()
        logger.info(f"DataLoader initialized: {file_path} (format: {self.format})")
    
    def _detect_format(self) -> str:
        """Detect file format from extension."""
        suffix = self.file_path.suffix.lower()
        
        if suffix in [".xlsx", ".xls"]:
            return "excel"
        elif suffix == ".json":
            return "json"
        elif suffix == ".csv":
            return "csv"
        else:
            raise ValueError(f"Unsupported file format: {suffix}")
    
    def load_text_cases(
        self, 
        sheet_name: str = "文本测试题",
        limit: Optional[int] = None,
    ) -> List[TestCase]:
        """
        Load text test cases.
        
        Args:
            sheet_name: Sheet name for Excel files
            limit: Maximum number of cases to load
            
        Returns:
            List of TestCase objects
        """
        if self.format == "excel":
            return self._load_excel(sheet_name, ContentType.TEXT, limit)
        elif self.format == "json":
            return self._load_json(ContentType.TEXT, limit)
        elif self.format == "csv":
            return self._load_csv(ContentType.TEXT, limit)
    
    def load_image_cases(
        self,
        sheet_name: str = "图片测试题",
        limit: Optional[int] = None,
    ) -> List[TestCase]:
        """
        Load image test cases.
        
        Args:
            sheet_name: Sheet name for Excel files
            limit: Maximum number of cases to load
            
        Returns:
            List of TestCase objects
        """
        if self.format == "excel":
            return self._load_excel(sheet_name, ContentType.IMAGE, limit)
        elif self.format == "json":
            return self._load_json(ContentType.IMAGE, limit)
        elif self.format == "csv":
            return self._load_csv(ContentType.IMAGE, limit)
    
    def _load_excel(
        self, 
        sheet_name: str, 
        content_type: ContentType,
        limit: Optional[int] = None,
    ) -> List[TestCase]:
        """Load test cases from Excel file."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for Excel files. Install with: pip install openpyxl")
        
        cases = []
        workbook = openpyxl.load_workbook(self.file_path, read_only=True)
        
        try:
            if sheet_name not in workbook.sheetnames:
                logger.warning(f"Sheet '{sheet_name}' not found. Available: {workbook.sheetnames}")
                return cases
            
            sheet = workbook[sheet_name]
            
            # Expected columns: 类型(黑/白), 序号, 内容, 预期风险
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
                if not row or not row[0]:  # Skip empty rows
                    continue
                
                if limit and len(cases) >= limit:
                    break
                
                case = TestCase(
                    id=str(row[1]) if len(row) > 1 and row[1] else f"{content_type.value}_{row_idx}",
                    content=str(row[2]) if len(row) > 2 and row[2] else "",
                    content_type=content_type,
                    expected_risk=str(row[3]) if len(row) > 3 and row[3] else "正常",
                    category=str(row[0]) if row[0] else "",
                    metadata={"row_number": row_idx + 1},
                )
                
                if case.content:  # Only add cases with content
                    cases.append(case)
            
            logger.info(f"Loaded {len(cases)} {content_type.value} cases from Excel")
            
        finally:
            workbook.close()
        
        return cases
    
    def _load_json(
        self, 
        content_type: ContentType,
        limit: Optional[int] = None,
    ) -> List[TestCase]:
        """Load test cases from JSON file."""
        cases = []
        
        with open(self.file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        
        # Support both array and object with 'text'/'image' keys
        if isinstance(data, list):
            items = data
        elif isinstance(data, dict):
            type_key = content_type.value
            items = data.get(type_key, data.get(f"{type_key}_cases", []))
        else:
            items = []
        
        for idx, item in enumerate(items):
            if limit and len(cases) >= limit:
                break
            
            case = TestCase(
                id=str(item.get("id", f"{content_type.value}_{idx}")),
                content=str(item.get("content", "")),
                content_type=content_type,
                expected_risk=str(item.get("expected_risk", item.get("label", "正常"))),
                category=str(item.get("category", item.get("type", ""))),
                metadata=item.get("metadata", {}),
            )
            
            if case.content:
                cases.append(case)
        
        logger.info(f"Loaded {len(cases)} {content_type.value} cases from JSON")
        return cases
    
    def _load_csv(
        self, 
        content_type: ContentType,
        limit: Optional[int] = None,
    ) -> List[TestCase]:
        """Load test cases from CSV file."""
        cases = []
        
        with open(self.file_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            
            for idx, row in enumerate(reader):
                if limit and len(cases) >= limit:
                    break
                
                case = TestCase(
                    id=str(row.get("id", row.get("序号", f"{content_type.value}_{idx}"))),
                    content=str(row.get("content", row.get("内容", ""))),
                    content_type=content_type,
                    expected_risk=str(row.get("expected_risk", row.get("预期风险", "正常"))),
                    category=str(row.get("category", row.get("类型", ""))),
                )
                
                if case.content:
                    cases.append(case)
        
        logger.info(f"Loaded {len(cases)} {content_type.value} cases from CSV")
        return cases
    
    def iter_cases(
        self, 
        content_type: ContentType,
        batch_size: int = 100,
    ) -> Iterator[List[TestCase]]:
        """
        Iterate over test cases in batches.
        
        Args:
            content_type: Type of content to load
            batch_size: Number of cases per batch
            
        Yields:
            Batches of TestCase objects
        """
        if content_type == ContentType.TEXT:
            all_cases = self.load_text_cases()
        else:
            all_cases = self.load_image_cases()
        
        for i in range(0, len(all_cases), batch_size):
            yield all_cases[i:i + batch_size]


def create_sample_data(output_path: str, format: str = "json") -> None:
    """
    Create sample test data file for reference.
    
    Args:
        output_path: Path to save the sample data
        format: Output format ('json' or 'csv')
    """
    sample_data = {
        "text": [
            {"id": "text_001", "content": "这是一条正常的文本内容", "expected_risk": "正常", "category": "白样本"},
            {"id": "text_002", "content": "敏感内容示例...", "expected_risk": "涉政", "category": "黑样本"},
            {"id": "text_003", "content": "广告推销内容...", "expected_risk": "广告", "category": "黑样本"},
        ],
        "image": [
            {"id": "img_001", "content": "https://example.com/normal.jpg", "expected_risk": "正常", "category": "白样本"},
            {"id": "img_002", "content": "https://example.com/sensitive.jpg", "expected_risk": "色情", "category": "黑样本"},
        ]
    }
    
    output_file = Path(output_path)
    
    if format == "json":
        with open(output_file, "w", encoding="utf-8") as f:
            json.dump(sample_data, f, ensure_ascii=False, indent=2)
    elif format == "csv":
        # Write text cases
        text_file = output_file.with_suffix(".text.csv")
        with open(text_file, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=["id", "content", "expected_risk", "category"])
            writer.writeheader()
            writer.writerows(sample_data["text"])
        
        # Write image cases
        image_file = output_file.with_suffix(".image.csv")
        with open(image_file, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=["id", "content", "expected_risk", "category"])
            writer.writeheader()
            writer.writerows(sample_data["image"])
    
    logger.info(f"Sample data created: {output_path}")
